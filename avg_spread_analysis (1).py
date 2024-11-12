import pandas as pd
from datetime import datetime, timedelta, date
from os.path import join, exists
import time
import sys
from multiprocessing import Pool
from openpyxl import load_workbook

sys.path.append('../Centralised Repo/Data Retrieval and APIs')
from xignites_optionsplay import retrieve_options_chain, APIError, retrieve_earnings, fetch_only_tickers, fundamentals_factset
from postgresql_interface import recorded_sql_queries, sql_querier
sys.path.append('../Centralised Repo/Commonly Used Functions')
from user_defined_exceptions import TickerFetchError


def pull_options_chain(tickers):
    def DiffSpotFilter(frame):
        # function to find the 4 least ITM puts and calls (least = closest to ATM)
        calls = frame[frame.Type == 'Call'].sort_values('DiffSpot')
        calls = calls[calls.DiffSpot <= 0].tail(4)

        puts = frame[frame.Type == 'Put'].sort_values('DiffSpot')
        puts = puts[puts.DiffSpot >= 0].head(4)

        return calls.append(puts, ignore_index=True).reset_index(drop=True)

    # calculate the date of the third friday of next month. bit convoluted but pretty fast.
    today = date.today()
    if today.month == 11:
        first_next_month = datetime(today.year + int((10 + 1) / 12), 12, 1)
    else:
        first_next_month = datetime(today.year + int((today.month + 1) / 12), (today.month + 1) % 12, 1)
    third_friday = first_next_month + timedelta(4 - first_next_month.weekday() + (0 if first_next_month.weekday() <= 4
                                                                                  else 7)) + timedelta(14)
    from_date = (third_friday - timedelta(2))
    to_date = (third_friday + timedelta(2))

    # retrieve options chain specifically for next month for given ticker
    options_chain = pd.DataFrame()
    for ticker in tickers:
        try:
            options_chain = options_chain.append(retrieve_options_chain(ticker, str(third_friday.month), str(third_friday.year)))
        except APIError:
            print('Problem retrieving options chain for {}'.format(ticker))
            continue

    # filter option chains to expiries within exactly +/-2 days from the third friday date calculated previously
    options_chain = options_chain[(options_chain.ExpirationDate >= from_date) & (options_chain.ExpirationDate <= to_date)
                                  & (options_chain.Ask != 0)]
    # sum open interest by symbol for use later
    open_interests = options_chain.groupby('Symbol')['OpenInterest'].sum().reset_index()

    # calculate DiffSpot as (Strike - Spot)
    options_chain.loc[:, 'DiffSpot'] = options_chain.StrikePrice - options_chain.Spot

    # do further filtering on symbols to find only the 4 least ITM puts and calls for each Symbol. then merge the
    # resultant table with the open interest sum table calculated before (merge on the Symbol column)
    options_chain = options_chain.groupby('Symbol').apply(DiffSpotFilter).reset_index(drop=True).drop('OpenInterest', axis=1)\
        .merge(open_interests, on='Symbol')

    return options_chain


def calc_avg_spread(options_chain):
    # calculates and averages SpreadMid and SpreadSpot for the entire options chain

    # calculate SpreadMid and SpreadSpot
    options_chain['SpreadMid'] = (options_chain.Ask - options_chain.Bid) / options_chain[['Bid', 'Ask']].mean(axis=1)
    options_chain['SpreadSpot'] = (options_chain.Ask - options_chain.Bid) / options_chain.Spot

    # take mean of SpreadMid, SpreadSpot and OpenInterest per Symbol and then filter that frame to get mean SpreadMid
    # ang SpreadSpot which are positive.
    avg_spread_frame = options_chain.groupby('Symbol').mean()[['Spot', 'SpreadMid', 'SpreadSpot', 'OpenInterest']].reset_index()
    avg_spread_frame = avg_spread_frame[(avg_spread_frame.SpreadMid > 0) & (avg_spread_frame.SpreadSpot > 0)]

    return avg_spread_frame


def do_everything(tickers):
    # function to perform all option chain specific tasks in one place

    def liquidity_rank(data_row):
        # function to calculate Liquidity Rank based on the values of SpreadSpot and OpenInterest
        if data_row['SpreadSpot'] < 0.004 and data_row['OpenInterest'] > 1000:
            return '1 (Very Liquid)'
        elif (data_row['SpreadSpot'] < 0.006 and data_row['OpenInterest'] > 1000) or (data_row['SpreadSpot'] < 0.004 and
                                                                                      data_row['OpenInterest'] < 1000):
            return '2 (Somewhat Liquid)'
        else:
            return '3 (Not Very Liquid)'

    # since this function is called in parallel, each instance can pull all the option chains of the tickers passed to
    # it in one go. If it turns out that all the tickers together couldn't produce even one proper option chain between
    # them, then just return an empty dataframe.
    options_chain = pull_options_chain(tickers)
    if options_chain.empty:
        return pd.DataFrame()

    # calculate the average spread columns for all options (see function description for more details)
    avg_spreads = calc_avg_spread(options_chain)
    combined_frame = avg_spreads
    # ^^ this rename is useless and was left in to keep compatibility with old code that dwindled further and further
    # until it was only the single line above

    # Insert column with liquidity rank calculated to just before the SpreadMid column
    combined_frame.insert(combined_frame.columns.get_loc('SpreadMid')-1, 'Liquidity',
                          combined_frame.loc[:, ['SpreadSpot', 'OpenInterest']].apply(liquidity_rank, axis=1))
    return combined_frame


def create_earnings_column(earnings_frame):
    # function to create earnings column in specific format

    # wherever earnings is not empty, write out new columns as 'AM' or 'PM' depending on whether earnings are released
    # before or after market, followed by the earnings date, followed by the days left to earnings inside parantheses.
    non_null_earnings = ~pd.isnull(earnings_frame['Earnings Date'])
    earnings_frame.loc[non_null_earnings, 'Earnings Date'] = earnings_frame.loc[non_null_earnings, 'TimeType'].\
        map({'BeforeMarket': 'AM', 'AfterMarket': 'PM'}).fillna('--').astype(str) + ' ' + \
        pd.to_datetime(earnings_frame.loc[non_null_earnings, 'Earnings Date']).dt.strftime('%m/%d/%Y') + ' (' + \
        (earnings_frame.loc[non_null_earnings, 'Earnings Date'] - date.today()).dt.days.astype(str) + ')'
    return earnings_frame


def write_to_file(filename, combined_frame):
    writer = pd.ExcelWriter(filename)
    write_frame = combined_frame

    write_frame.sort_values('Liquidity').to_excel(writer, index=False)
    book = writer.book
    sheet = writer.sheets['Sheet1']

    money_fmt = book.add_format({'num_format': '$#,##0.00'})
    perc_fmt = book.add_format({'num_format': '0%'})
    decimal_fmt = book.add_format({'num_format': '0.00'})
    sheet.set_column(2, 3, cell_format=perc_fmt)
    sheet.set_column(1, 1, cell_format=money_fmt)
    # sheet.set_column(12, 15, cell_format=decimal_fmt)

    for ind in range(len(write_frame.columns)):
        max_len = max(len(write_frame.columns[ind]), write_frame.iloc[:, ind].astype(str).map(len).max()) + 1
        sheet.set_column(ind, ind, max_len)

    sheet.autofilter(0, 0, write_frame.shape[0] - 1, write_frame.shape[1] - 1)
    writer.save(), writer.close()

    return combined_frame


def create_liquidity_file(combined_frame, filename):
    # function to create the liquidity tracker file that is simply a rearrangement of a select part of data from the
    # main liquidity file.

    writer = pd.ExcelWriter(filename, engine='openpyxl')
    if exists(filename):
        book = load_workbook(filename)
        writer.book = book
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
        sheet = book['Sheet1']
        for col in range(1, 5):
            for row in range(1, 4000):
                sheet.cell(row, col).value = None
    else:
        book = writer.book

    # sort by Liquidity and SpreadSpot columns and write rows with ranks '1 (Very Liquid)' and '2 (Somewhat Liquid)'
    # into the new file (but only select columns from these rows)
    combined_frame.sort_values(['Liquidity', 'SpreadSpot'], inplace=True)
    combined_frame = combined_frame[~combined_frame['Liquidity'].str.contains('3')][['Symbol', 'Liquidity', 'IV Rank',
                                                                                     'Earnings Date']]
    combined_frame.to_excel(writer, sheet_name='Sheet1', index=False)

    sheet = book['Sheet1']
    sheet.cell(1, 5).value = 'Last Updated: '
    sheet.cell(1, 6).value = date.today().strftime('%m/%d/%Y')

    for ind in range(len(combined_frame.columns)):
        max_len = max(len(combined_frame.columns[ind]), combined_frame.iloc[:, ind].astype(str).map(len).max()) + 1
        sheet.column_dimensions[sheet.cell(1, ind+1).column_letter].width = max_len

    sheet.column_dimensions[sheet.cell(1, 5).column_letter].width = len(sheet.cell(1, 5).value)
    sheet.column_dimensions[sheet.cell(1, 6).column_letter].width = len(sheet.cell(1, 6).value)

    perc_fmt1 = '0%'
    for col in [combined_frame.columns.get_loc(c)+1 for c in ['IV Rank']]:
        for row in range(2, combined_frame.shape[0] + 2):
            sheet.cell(row, col).number_format = perc_fmt1

    writer.save(), writer.close()


def read_sql_data():
    # use the recorded SQL query along with all the defaults of the sql_querier function to retrieve IV rank and
    # percentile for the latest date
    vol_data = sql_querier(recorded_sql_queries['latest_iv_rank_perc'])

    # divide numbers by 100 to match up rest of the code and rename column names for the rest of the code to work
    vol_data.loc[:, ['IV rank', 'IV percentile']] /= 100
    vol_data.rename(columns={'Ticker': 'Symbol', 'IV percentile': 'IV Percentile', 'IV rank': 'IV Rank'}, inplace=True)
    return vol_data


def testing():
    # testing module. this should ideally be its own code on a separate branch. but for now, simply replace "main()" on
    # the "if ___name__ ..." control to "testing()"
    tickers = fetch_only_tickers('OPRA')[:50]
    earnings = []
    for ticker in tickers:
        earnings.append(retrieve_earnings(ticker))
    earnings = pd.concat(earnings)
    earnings = create_earnings_column(earnings)
    pass


def main():
    start_time = time.time()

    # if we can't fetch tickers, just quit entirely.
    try:
        tickers = fetch_only_tickers('OPRA')
    except TickerFetchError:
        print('There was a problem fetching tickers in the very first step. Program is exiting.')
        exit(-1)

    # there are 2 outputs from this code
    # 1) the main liquidity report
    # 2) the liquidity tracker which is a filtered down version of the liquidity report
    output_file = join('C:\\Users\\op-quant-one\\OneDrive - OptionsPlay\\Shared with Everyone', 'Option Liquidity - US.xlsx')
    liquidity_file = join('C:\\Users\\op-quant-one\\OneDrive - OptionsPlay\\Shared with Everyone',
                          'OptionsPlay Options Liquidity Tracker - US.xlsx')

    # start a process pool and then throw tasks into it that can be executed in parallel. python wrappers are hella
    # clean so it makes the whole concept trivially simple.
    all_proc_pool = Pool(processes=10, maxtasksperchild=10)
    earnings = all_proc_pool.map_async(retrieve_earnings, tickers)  # earnings task
    fundamentals = all_proc_pool.apply_async(fundamentals_factset, [tickers])  # fundamentals task
    vol_data = all_proc_pool.apply_async(read_sql_data, ())  # vol data retrieval from SQL database

    # options chain retrieval wrapper function. pass 100 tickers to it at a time for good results
    options_chain = all_proc_pool.map_async(do_everything, [tickers[i:i+100] for i in range(0, len(tickers), 100)])
    all_proc_pool.close(), all_proc_pool.join()

    # merge all outputs together (mostly on the Symbol columns) and then do some column renaming etc to prettify the
    # output file generated. also run a function to generate earnings column in a specific format from earnings frame.
    combined_frame = pd.merge(pd.concat(options_chain.get()), create_earnings_column(pd.concat(earnings.get())),
                              how='left', on='Symbol')
    combined_frame = combined_frame.merge(vol_data.get(), how='left', on='Symbol')
    combined_frame = combined_frame.merge(fundamentals.get().drop('MarketCap', axis=1), how='left', on='Symbol')
    combined_frame.rename(columns={'Spot': 'Price', 'OpenInterest': 'Open Interest'}, inplace=True)
    column_order = ['Symbol', 'Price', 'IV Rank', 'IV Percentile', 'Liquidity', 'Earnings Date', 'Open Interest'] + \
                   ['Sector', 'Subsector']

    # separate function to write the prettified output. will probably be replaced in the future with a
    # centralised version
    write_to_file(output_file, combined_frame[column_order])

    # create separate liquidity tracker file
    create_liquidity_file(combined_frame, liquidity_file)

    print('Code run in %.2f seconds.' % (time.time() - start_time))


if __name__ == '__main__':
    main()
    # testing()
